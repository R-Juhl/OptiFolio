from flask import Flask, jsonify, request, Response, send_file, send_from_directory
from flask_cors import CORS
import yfinance as yf # for historical stock data
import numpy as np # for numerical operations
from scipy.optimize import minimize # for optimization
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
import openai
import os

#for debugging:
#print("Current working directory:", os.getcwd())
#print(app.config)

api_key = os.environ.get("OPENAI_API_KEY")

static_folder_path = os.path.join(os.getcwd(), 'optifolio', 'build')
app = Flask(__name__, static_folder=static_folder_path, static_url_path='/')

cors_origin = os.environ.get("CORS_ORIGIN", "http://localhost:3000")
CORS(app, origins=[cors_origin])

@app.route('/test')
def test():
    return 'Test route is working!'

@app.route('/test-static')
def test_static():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

def fetch_stock_data(stocks, start_date, end_date):
    data = yf.download(stocks, start=start_date, end=end_date)['Adj Close']
    return data

def compute_portfolio_metrics(data):
    returns = data.pct_change().dropna()
    expected_returns = returns.mean() * 252
    covariance_matrix = returns.cov() * 252
    correlation_matrix = returns.corr()
    return expected_returns, covariance_matrix, correlation_matrix

def compute_optimal_portfolio(expected_returns, covariance_matrix, target_return=None, short=False):
    num_assets = len(expected_returns)
    args = (expected_returns, covariance_matrix)
    constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
    
    if target_return:
        constraints = (
            {'type': 'eq', 'fun': lambda x: np.sum(x) - 1},
            {'type': 'eq', 'fun': lambda x: target_return - np.dot(x, expected_returns)}
        )
    
    bound = (1 if short else 0, 1)
    bounds = tuple(bound for asset in range(num_assets))
    
    result = minimize(portfolio_performance, num_assets*[1./num_assets,], args=args, method='SLSQP', bounds=bounds, constraints=constraints)
    return result.x

def portfolio_performance(weights, expected_returns, covariance_matrix):
    port_return = np.dot(weights, expected_returns)
    port_volatility = np.sqrt(np.dot(weights.T, np.dot(covariance_matrix, weights)))
    sharpe_ratio = port_return / port_volatility
    return -sharpe_ratio  # negative for maximization

@app.route('/api/compute-portfolio', methods=['POST'])
def compute_portfolio():
    stocks = request.json['stocks']
    start_date = request.json['startDate']
    end_date = request.json['endDate']
    risk_level = request.json['riskLevel']  # Get risk level from the request

    # Convert risk_level to an integer
    try:
        risk_level = int(risk_level)
    except ValueError:
        return jsonify({
            "message": "Invalid risk level. Risk level should be a number.",
        }), 400  # Return an error status if risk_level is not a number

    # If only one stock is selected, return a message
    if len(stocks) == 1:
        return jsonify({
            "message": "Portfolio calculations require more than one stock.",
        }), 400  # Return an error status

    try:
        data = fetch_stock_data(stocks, start_date, end_date)
    except Exception as e:
        return jsonify({
            "message": f"Error fetching data for stocks: {e}",
            "error": "Failed to retrieve historical data."
        }), 400  # Return an error status

    expected_returns, covariance_matrix, _ = compute_portfolio_metrics(data)
    min_return = min(expected_returns)
    max_return = max(expected_returns)
    target_returns = np.linspace(min_return, max_return, 12)  # 12 points on the frontier

    frontier = []
    for target in target_returns:
        weights = compute_optimal_portfolio(expected_returns, covariance_matrix, target)
        port_return = np.dot(weights, expected_returns)
        port_volatility = np.sqrt(np.dot(weights.T, np.dot(covariance_matrix, weights)))
        frontier.append({"return": port_return, "volatility": port_volatility})

    # Tangency Portfolio (highest Sharpe ratio)
    tangency_portfolio = max(frontier, key=lambda x: x["return"] / x["volatility"])
    # Individual Stock Data
    individual_stocks = [{"name": stock, "return": expected_returns[stock], "volatility": (covariance_matrix[stock][stock])**0.5} for stock in stocks]
    # Minimum Variance Portfolio
    min_variance_portfolio = min(frontier, key=lambda x: x["volatility"])
    
    # Compute the optimal portfolio based on the risk level
    if risk_level <= 5:
        # Scale between minimum variance and tangency
        scale = (risk_level - 1) / 4.0  # Scale from 0 to 1 as risk level goes from 1 to 5
        optimal_return = min_variance_portfolio['return'] + (tangency_portfolio['return'] - min_variance_portfolio['return']) * scale
    else:
        # Scale between tangency and maximum return
        scale = (risk_level - 5) / 5.0  # Scale from 0 to 1 as risk level goes from 5 to 10
        optimal_return = tangency_portfolio['return'] + (max_return - tangency_portfolio['return']) * scale

    # Find the closest matching portfolio on the frontier
    optimal_portfolio = min(frontier, key=lambda x: abs(x['return'] - optimal_return))

    # Calculate the weights for the optimal portfolio again since we need them to match the optimal_return
    optimal_weights = compute_optimal_portfolio(expected_returns, covariance_matrix, optimal_portfolio['return'])

    return jsonify({
        "weights": optimal_weights.tolist(), 
        "frontier": frontier, 
        "tangency": tangency_portfolio,
        "stocks": individual_stocks,
        "min_variance": min_variance_portfolio,
        "optimalPortfolio": optimal_portfolio
    })

@app.route('/api/validate-stock', methods=['POST'])
def validate_stock():
    stock = request.json['stock']
    try:
        # fetching data for the last 7 days as validity check (because shortName did not work for some reason)
        data = yf.download(stock, period="7d")
        if data.empty:
            raise ValueError(f"No data available for {stock}")
        return jsonify({"valid": True, "message": f"{stock} is a valid stock."})
    except Exception as e:
        print(f"Error validating stock {stock}: {e}")
        return jsonify({"valid": False, "message": f"{stock} is not a valid stock."})

from datetime import datetime, timedelta

@app.route('/api/check-daterange', methods=['POST'])
def check_date_range():
    stocks = request.json['stocks']
    start_date = datetime.strptime(request.json['startDate'], '%Y-%m-%d')
    end_date = datetime.strptime(request.json['endDate'], '%Y-%m-%d')
    missing_data_stocks = []
    
    for stock in stocks:
        data = yf.download(stock, start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d'), progress=False)
        data_start_date = data.index[0]
        data_end_date = data.index[-1]
        
        if abs((data_start_date - start_date).days) > 7 or abs((data_end_date - end_date).days) > 7:
            missing_data_stocks.append(stock)
            
    if missing_data_stocks:
        message = f"Stocks {', '.join(missing_data_stocks)} do not have continuous historical data for the entire specified date range."
        return jsonify({"valid": False, "message": message})
    else:
        return jsonify({"valid": True, "message": ""})

last_call_time = None  # To keep track of the last time the API was called

@app.route('/api/chat-gpt', methods=['POST'])
def chat_gpt():
    global last_call_time
    query = request.json['query']
    
    # Check rate limiting
    now = datetime.now()
    if last_call_time and (now - last_call_time).total_seconds() < 30:
        return jsonify({"error": "You can only make a request every 30 seconds."}), 429
    
    openai.api_key = api_key
    
    user_message = request.json['query']
    context = "I want you to recommend/suggest publicly traded companies based on the message below. I want you to recommend/suggest multiple companies when relevant or requested and send them in a list, and always with the companies tickers following their respective names in parenthesis. The context is that the user is looking for suggestions of industries and/or companies to research further for possible investment purposes. This is the user message:"
    
    # Make the API call
    messages = [{"role": "system", "content": context}, {"role": "user", "content": user_message}]
    response = openai.ChatCompletion.create(model="gpt-4", messages=messages)
    
    last_call_time = datetime.now()
    return jsonify({"response": response.choices[0].message["content"].strip()})

@app.route('/api/get-stock-prices', methods=['POST'])
def get_stock_prices():
    print("Endpoint hit!")
    stocks = request.json['stocks']
    print("Fetching data for stocks:", stocks)
    try:
        data = yf.download(stocks, period="1d")['Adj Close'].iloc[-1]  # Last row for current prices
        print("Data fetched:", data.to_dict())
        return jsonify({"prices": data.to_dict()})
    except Exception as e:
        print("Error:", e)
        return jsonify({
            "message": f"Error fetching data for stocks: {e}",
            "error": "Failed to retrieve stock prices."
        }), 400

@app.route('/api/generate-excel', methods=['POST'])
def generate_excel():
    data = request.json
    stocks = data['stocks']
    target_percentages = data['targetPercentages']
    current_shares = data['currentShares']
    current_prices = data['currentPrices']
    shares_to_buy = data['sharesToBuy']
    surplus = data['surplus']
    fee = data['fee']
    monthly_investment_plan = data['monthlyInvestmentPlan']
    current_month = datetime.now().month

    # Load the Excel template
    template_path = 'plan/OptiFolio_Template.xlsm'
    wb = load_workbook(filename=template_path, keep_vba=True)
    ws = wb["Investment Plan"]

    # Styles
    orange_fill = PatternFill(start_color='ffb560', end_color='ffb560', fill_type='solid')
    light_fill = PatternFill(start_color='6B667B', end_color='6B667B', fill_type='solid')
    dark_fill = PatternFill(start_color='282c34', end_color='282c34', fill_type='solid')
    header_fill = PatternFill(start_color='4D495C', end_color='4D495C', fill_type='solid')
    red_fill = PatternFill(start_color='F27676', end_color='F27676', fill_type='solid')
    
    bold_orange_font = Font(bold=True, color="ffb560")
    bold_font = Font(bold=True)
    white_font = Font(color="EDEDED")
    bold_white_font = Font(bold=True, color="EDEDED")

    # Add table headers for initial investments based on capital
    headers = ["Stocks", "Current Price", "Target Allocation (%)", "Current Allocation (%)", "Current shares", "", "Monthly Surplus", "Fees (per transaction)"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_num).value = header

    # Write initial data
    for i, stock in enumerate(stocks, start=3):
        current_price = round(float(current_prices[i-3]), 1)
        current_share = float(current_shares[i-3])

        # Apply the light_fill style to all cells in the row except for specific columns
        for j in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=j)
            if cell.column_letter not in ['']:  # Option to exclude columns from formatting
                cell.fill = light_fill
                cell.font = bold_white_font

            # Now, populate the cell with value
            if j == 1:
                cell.value = stock
            elif j == 2:
                cell.value = current_price
            elif j == 3:
                cell.value = round(target_percentages[i-3], 1)
            elif j == 5:
                cell.value = current_share

    # Construct the sum formula for the "Current Allocation (%)"
    sum_formula = f"SUMPRODUCT(E3:E{len(stocks)+2}, B3:B{len(stocks)+2})"
    
    # Now set the formula for each "Current Allocation (%)" cell
    for i in range(3, len(stocks) + 3):
        allocation_formula = f"=ROUND((E{i}*B{i})/({sum_formula})*100, 1)"
        ws.cell(row=i, column=4).value = allocation_formula

    # Set surplus and fee in cell as values
    ws['G3'] = float(surplus)
    ws['H3'] = float(fee)
    ws['G4'] = len(stocks)

    # Apply styles
    ws.cell(row=3, column=7).fill = orange_fill
    ws.cell(row=3, column=8).fill = orange_fill
    ws.cell(row=4, column=7).fill = red_fill # todo: change this to instead hide the text (len(stocks))

    for i in range(3, len(stocks) + 3):
        ws.cell(row=i, column=5).fill = orange_fill

    # Define ranges for the VBA function
    stocks_range = f"A3:A{len(stocks)+2}"
    prices_range = f"B3:B{len(stocks)+2}"
    weights_range = f"C3:C{len(stocks)+2}"
    current_shares_range = f"E3:E{len(stocks)+2}"

    # color style for offset/spacing
    row_offset = len(stocks) + 3
    for col in range(1, 9):  # Columns A to H
        ws.cell(row=row_offset, column=col).fill = dark_fill
    row_offset += 1
    for col in range(1, 9):  # Columns A to H
        ws.cell(row=row_offset, column=col).fill = dark_fill
    row_offset += 1
    
    # Write initial buy period
    month_name = datetime(1900, (current_month - 1) % 12 + 1, 1).strftime('%B')
    for col in range(1, 9):  # This will fill cells from A to H
        cell = ws.cell(row=row_offset, column=col)
        cell.fill = header_fill
        cell.font = bold_orange_font
        if col == 1:
            cell.value = f"Initial investment ({month_name})"

    row_offset += 1

    for col in range(1, 9):  # This will fill cells from A to H
        cell = ws.cell(row=row_offset, column=col)
        cell.fill = header_fill
        cell.font = bold_white_font
        if col == 1:
            cell.value = "Stocks"
        elif col == 2:
            cell.value = "Shares to Buy"

    row_offset += 1

    for i, stock in enumerate(stocks, start=row_offset):
        ws.cell(row=i, column=1).value = stock
        ws.cell(row=i, column=2).value = shares_to_buy[i - row_offset]
        for col in range(1, 9):  # This will fill cells from A to H
            cell = ws.cell(row=i, column=col)
            cell.fill = light_fill
            cell.font = white_font

    # color style for offset/spacing
    row_offset += len(stocks)
    for col in range(1, 9):  # Columns A to H
        ws.cell(row=row_offset, column=col).fill = dark_fill
    row_offset += 1
    for col in range(1, 9):  # Columns A to H
        ws.cell(row=row_offset, column=col).fill = dark_fill
    row_offset += 1

    # Start writing monthly plan
    for period, monthly_data in enumerate(monthly_investment_plan, start=1):
        month_name = datetime(1900, (current_month + period - 1) % 12 + 1, 1).strftime('%B')
        cell = ws.cell(row=row_offset, column=1)
        cell.value = f"Month {period} ({month_name})"
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row_offset, column=col)
            cell.fill = header_fill
            cell.font = bold_orange_font
        row_offset += 1

        # Insert headers for the period
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row_offset, column=col)
            cell.fill = header_fill
            cell.font = bold_white_font
            if col == 1:
                cell.value = "Stocks"
            elif col == 2:
                cell.value = "Shares to Buy"
        row_offset += 1

        # Call the VBA function to get shares to buy for the period
        for i, stock in enumerate(stocks, start=row_offset):
            ws.cell(row=i, column=1).value = stock
            ws.cell(row=i, column=2).value = f'=INDEX(CalculateSharesToBuy({stocks_range}, {weights_range}, {prices_range}, {current_shares_range}, {surplus}, {fee}), {i-row_offset+1})'
            for col in range(1, 9):  # Columns A to H
                cell = ws.cell(row=i, column=col)
                cell.fill = light_fill
                cell.font = white_font
    
        row_offset += len(stocks) + 0
        for col in range(1, 9):  # Columns A to H
            ws.cell(row=row_offset, column=col).fill = dark_fill
        row_offset += 1
        for col in range(1, 9):  # Columns A to H
            ws.cell(row=row_offset, column=col).fill = dark_fill
        row_offset += 1

    # Save the changes to a BytesIO object to send as a response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(output.getvalue(), mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')
    response.headers['Content-Disposition'] = 'attachment; filename=OptiFolio Investment Plan.xlsm'
    return response

if __name__ == '__main__':
    app.run(debug=False)